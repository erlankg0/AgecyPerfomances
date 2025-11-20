import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# Конфигурация
FILE_PATH = "input-data/agency/agency.xlsx"
OUTPUT_PATH = "output-data/agency_group_sales.xlsx"

print("🚀 Начало обработки данных...")

# ============================================
# 1. ЗАГРУЗКА И ОЧИСТКА ДАННЫХ
# ============================================
print("📥 Загрузка данных...")
raw = pd.read_excel(FILE_PATH, header=None)
header_row = raw[raw[0] == "Agency"].index[0]
df = pd.read_excel(FILE_PATH, header=header_row)

# Удаляем пустые строки и колонки
df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")

# Нормализация колонок
df.columns = (
    df.columns.astype(str)
    .str.strip()
    .str.replace('\n', '_', regex=True)
    .str.replace('_x000a_', '_', regex=True)
    .str.replace(' ', '_', regex=True)
    .str.replace('__+', '_', regex=True)
    .str.replace(r'[^\w_%\.]', '', regex=True)
    .str.lower()
)

print(f"Исходных строк после очистки: {len(df)}")

# ============================================
# 2. ИЗВЛЕЧЕНИЕ МЕСЯЦЕВ
# ============================================
print("📅 Извлечение месяцев...")
month_pattern = r'\d{2}-[^\d]+'
month_rows = df['agency'].astype(str).str.match(month_pattern, na=False)

df['month'] = pd.NA
current_month = None

for i in df.index:
    if month_rows[i]:
        month_str = df.at[i, 'agency']
        parts = str(month_str).split('-')
        if len(parts) >= 2:
            current_month = parts[1].strip()
    if current_month:
        df.at[i, 'month'] = current_month

# Удаляем строки с заголовками месяцев
df = df[~month_rows].copy()
df = df.reset_index(drop=True)

print(f"После удаления заголовков месяцев: {len(df)}")

# ============================================
# 3. ОПРЕДЕЛЕНИЕ РЫНКОВ И УДАЛЕНИЕ СЛУЖЕБНЫХ СТРОК
# ============================================
print("🌍 Определение рынков...")

# Словарь для сокращенных названий рынков
market_names_map = {
    'CIS_COMMONWEALTH OF INDEPENDENT STATES': 'CIS',
    'DOMESTIC_DOMESTIC': 'DOMESTIC',
    'EUROPE_EUROPE MARKET': 'EUROPE',
    'MIDDLEEAST_MIDDLE EAST MARKET': 'ORTA DOĞU',
    'OTHER_OTHER MARKETS': 'OTHER',
    'FAR EASTERN_UZAK DOGU ULKERI': 'FAR EAST',
    'FAR EASTER_UZAK DOGU ULKERI': 'FAR EAST'
}

# Служебные строки, которые нужно удалить
service_rows_patterns = ['TOTAL', 'UK_UNITED KINGDOM'] + list(market_names_map.keys())

df['market'] = pd.NA
df['is_service_row'] = False
current_market = None

for i in df.index:
    agency_val = str(df.at[i, 'agency']).strip()
    agency_upper = agency_val.upper()

    # Проверяем, является ли строка заголовком рынка
    is_market_header = False
    for market_key in market_names_map.keys():
        if agency_upper == market_key.upper():
            current_market = market_names_map[market_key]
            is_market_header = True
            df.at[i, 'is_service_row'] = True
            break

    # Проверяем, является ли строка TOTAL (точное совпадение или содержит TOTAL)
    if agency_upper == 'TOTAL' or 'TOTAL' in agency_upper:
        df.at[i, 'is_service_row'] = True

    # Проверяем UK_UNITED KINGDOM
    if 'UK_UNITED KINGDOM' in agency_upper or agency_upper == 'UK_UNITED KINGDOM':
        df.at[i, 'is_service_row'] = True

    # Если это не служебная строка, присваиваем текущий рынок
    if not df.at[i, 'is_service_row']:
        df.at[i, 'market'] = current_market

# Удаляем все служебные строки ОДНИМ РАЗОМ
df_clean = df[df['is_service_row'] == False].copy()
df_clean = df_clean.drop(columns=['is_service_row'])
df_clean = df_clean.reset_index(drop=True)

print(f"После удаления служебных строк: {len(df_clean)}")
print(f"Найдено рынков: {df_clean['market'].nunique()}")
print(f"Рынки: {sorted(df_clean['market'].dropna().unique())}")

# ============================================
# 4. ПРЕОБРАЗОВАНИЕ ЧИСЛОВЫХ КОЛОНОК
# ============================================
print("🔢 Обработка числовых данных...")

numeric_cols = ['arrival_room', 'night_room', 'night_paidpax', 'eur_revenue', 'eur_avg_perpaidpax']
for col in numeric_cols:
    if col in df_clean.columns:
        df_clean[col] = pd.to_numeric(
            df_clean[col].astype(str).str.replace(',', '.').str.replace(' ', ''),
            errors='coerce'
        )

# ============================================
# 5. ГРУППИРОВКА АГЕНТСТВ
# ============================================
print("🏢 Группировка агентств...")

agency_group_rules = {
    'Anex Tour': ['ANEX-'],
    'AKAY TOUR': ['AKAY-'],
    'ARELES (EUROPEHOL)': ['ARELES-'],
    'BEDSOPIA / PRIME TRAVEL': ['BEDSOPIA'],
    'BOOKING.COM': ['BOOKING.COM'],
    'COMP': ['GM', 'COMP 3', 'SALES', 'KONSER', 'ATG', 'PANDEMI'],
    'CORENDON': ['CORENDON'],
    'DESTINATION SERVICES': ['DESTINATION-'],
    'ETS': ['ETS'],
    'EUROPE HOLIDAY': ['EUHOLIDAY-'],
    'FIBULA TRAVEL': ['FIBULA-'],
    'FIT TURIZM': ['FIT HOL-', 'FIT'],
    'GROUP': ['GROUP-'],
    'HOTELBEDS': ['HOTELBEDS-'],
    'HOUSE USE': ['HOUSE USE'],
    'INDIVIDUAL': ['INDIVIDUAL-'],
    'ITS': ['ITS-'],
    'KALANIT TOUR': ['KALANIT-'],
    'KEYF TRAVEL': ['KEYF TRAVEL-', 'SUNQUEST-'],
    'KILIT GLOBAL': ['KILIT-'],
    'MEETING POINT': ['FTI-'],
    'MOTUS': ['MOTUS-'],
    'ODEON TOUR': ['ODEON-'],
    'PASSO TOUR': ['PASSO-'],
    'PENINSULA': ['PENINSULA-'],
    'PGM HOLIDAY': ['PGM HOLIDAY-'],
    'RUSTAR': ['RUSTAR'],
    'SETUR': ['SETUR'],
    'SONAR TOUR': ['SONAR-'],
    'SUMMER TOUR': ['SUMMER-'],
    'TATILBUDUR': ['TATILBUDUR'],
    'WEB': ['WEB-'],
    'ZEYDE TURIZM': ['ZEYDE TURIZM']
}

def map_agency_group_partial(agency_name):
    agency_upper = str(agency_name).upper()
    for group_name, patterns in agency_group_rules.items():
        for pattern in patterns:
            # Проверяем начало строки или после пробела
            if agency_upper.startswith(pattern.upper()) or f" {pattern.upper()}" in agency_upper:
                return group_name
    # Если ни одно правило не подошло, возвращаем SORSAT
    return 'SORSAT'

df_clean['agency_group'] = df_clean['agency'].apply(map_agency_group_partial)

print(f"Найдено групп агентств: {df_clean['agency_group'].nunique()}")

# ============================================
# 6. ФИНАЛЬНАЯ ОЧИСТКА
# ============================================
print("🧹 Финальная очистка...")

# Удаляем строки без месяца или рынка
df_clean = df_clean.dropna(subset=['month', 'market']).copy()
df_clean = df_clean.reset_index(drop=True)

print(f"После удаления строк без month/market: {len(df_clean)}")

# ============================================
# 7. ГРУППИРОВКА ДАННЫХ
# ============================================
print("📊 Группировка данных...")

group_cols = ['month', 'market', 'agency_group', 'agency']
df_grouped = df_clean.groupby(group_cols, dropna=False)[numeric_cols].sum(min_count=1).reset_index()

print(f"Итоговых записей после группировки: {len(df_grouped)}")

# Создаем дополнительные сводки
df_by_group = df_grouped.groupby(['month', 'agency_group'], dropna=False).agg({
    'arrival_room': 'sum',
    'night_room': 'sum',
    'eur_revenue': 'sum',
    'eur_avg_perpaidpax': 'mean'
}).reset_index().sort_values('eur_revenue', ascending=False)

df_by_market = df_grouped.groupby(['month', 'market'], dropna=False).agg({
    'arrival_room': 'sum',
    'night_room': 'sum',
    'eur_revenue': 'sum',
    'eur_avg_perpaidpax': 'mean'
}).reset_index().sort_values('eur_revenue', ascending=False)

# ============================================
# 8. ФОРМАТИРОВАНИЕ И ВИЗУАЛИЗАЦИЯ EXCEL
# ============================================
print("💾 Сохранение отчета с визуализацией...")

def format_worksheet(ws):
    """Форматирует worksheet с красивыми стилями"""
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Форматируем заголовки
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Форматируем данные
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center')

            # Форматируем числовые значения
            if isinstance(cell.value, (int, float)) and cell.value is not None:
                cell.number_format = '#,##0.00'

    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Закрепляем первую строку
    ws.freeze_panes = 'A2'

# Графики удалены по запросу пользователя

# Сохраняем с форматированием
with pd.ExcelWriter(OUTPUT_PATH, engine='openpyxl') as writer:
    # Записываем все листы
    df_grouped.to_excel(writer, sheet_name='Summary', index=False)
    df_by_group.to_excel(writer, sheet_name='By Agency Group', index=False)
    df_by_market.to_excel(writer, sheet_name='By Market', index=False)

    workbook = writer.book

    # Форматирование каждого листа
    for sheet_name in ['Summary', 'By Agency Group', 'By Market']:
        ws = workbook[sheet_name]
        format_worksheet(ws)

# ============================================
# 9. СТАТИСТИКА
# ============================================
print("\n" + "="*60)
print("✅ ОБРАБОТКА ЗАВЕРШЕНА!")
print("="*60)
print(f"📁 Файл сохранен: {OUTPUT_PATH}")
print(f"📋 Всего записей в Summary: {len(df_grouped)}")
print(f"📊 Агентских групп: {df_grouped['agency_group'].nunique()}")
print(f"🌍 Рынков: {df_grouped['market'].nunique()}")
print(f"📅 Месяцев: {df_grouped['month'].nunique()}")
print(f"💰 Общая выручка (EUR): {df_grouped['eur_revenue'].sum():,.2f}")
print(f"🏨 Всего комнат: {df_grouped['arrival_room'].sum():,.0f}")
print(f"🌙 Всего ночей: {df_grouped['night_room'].sum():,.0f}")
print("="*60)

# Показываем распределение по группам
print("\n📊 Топ-10 групп по выручке:")
top_groups = df_grouped.groupby('agency_group')['eur_revenue'].sum().sort_values(ascending=False).head(10)
for idx, (group, revenue) in enumerate(top_groups.items(), 1):
    print(f"{idx}. {group}: {revenue:,.2f} EUR")

print("\n🌍 Распределение по рынкам:")
market_dist = df_grouped.groupby('market')['eur_revenue'].sum().sort_values(ascending=False)
for market, revenue in market_dist.items():
    print(f"  {market}: {revenue:,.2f} EUR")