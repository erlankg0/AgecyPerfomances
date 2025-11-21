from PySide6.QtCore import QObject, Signal, Slot


class Worker(QObject):
    finished = Signal(str)
    error = Signal(str)
    log = Signal(str)

    def __init__(self, input_file, pairs_file, output_file):
        super().__init__()
        self.input_file = input_file
        self.pairs_file = pairs_file
        self.output_file = output_file

    @Slot()
    def run(self):
        try:
            import pandas as pd
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            import warnings
            warnings.filterwarnings('ignore')

            self.log.emit("🚀 Обработка началась...")

            FILE_PATH = self.input_file
            OUTPUT_PATH = self.output_file

            # ============================
            # 1. LOAD
            # ============================
            self.log.emit("📥 Okuma Excel...")
            raw = pd.read_excel(FILE_PATH, header=None)
            header_row = raw[raw[0] == "Agency"].index[0]
            df = pd.read_excel(FILE_PATH, header=header_row)

            df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")

            df.columns = (
                df.columns.astype(str)
                .str.strip()
                .str.replace('\n', '_', regex=True)
                .str.replace('_x000a_', '_', regex=True)
                .str.replace(' ', '_', regex=True)
                .str.replace('__+', '_', regex=True)
                .str.replace(r'[^\w_%\.]', '', regex=True)
                .str.lower()
            )
            print(df.columns)

            # ============================
            # 2. MONTH
            # ============================
            self.log.emit("📅 Ayları arama...")

            month_pattern = r'\d{2}-[^\d]+'
            month_rows = df['agency'].astype(str).str.match(month_pattern, na=False)

            df['month'] = pd.NA
            current_month = None

            for i in df.index:
                if month_rows[i]:
                    month_str = df.at[i, 'agency']
                    parts = str(month_str).split('-')
                    if len(parts) >= 2:
                        current_month = parts[1].strip()
                if current_month:
                    df.at[i, 'month'] = current_month

            df = df[~month_rows].reset_index(drop=True)

            # ============================
            # 3. MARKET DETECT
            # ============================
            self.log.emit("🌍 Pazarları  Arama...")

            market_names_map = {
                'CIS_COMMONWEALTH OF INDEPENDENT STATES': 'CIS',
                'DOMESTIC_DOMESTIC': 'DOMESTIC',
                'EUROPE_EUROPE MARKET': 'EUROPE',
                'MIDDLEEAST_MIDDLE EAST MARKET': 'ORTA DOĞU',
                'OTHER_OTHER MARKETS': 'OTHER',
                'FAR EASTERN_UZAK DOGU ULKERI': 'FAR EAST',
                'FAR EASTER_UZAK DOGU ULKERI': 'FAR EAST'
            }

            df['market'] = pd.NA
            df['is_service_row'] = False
            current_market = None

            for i in df.index:
                agency_val = str(df.at[i, 'agency']).strip()
                agency_upper = agency_val.upper()

                # заголовки рынков
                for mk in market_names_map.keys():
                    if agency_upper == mk.upper():
                        current_market = market_names_map[mk]
                        df.at[i, 'is_service_row'] = True

                # TOTAL
                if agency_upper == 'TOTAL' or 'TOTAL' in agency_upper:
                    df.at[i, 'is_service_row'] = True

                if 'UK_UNITED KINGDOM' in agency_upper:
                    df.at[i, 'is_service_row'] = True

                if not df.at[i, 'is_service_row']:
                    df.at[i, 'market'] = current_market

            df_clean = df[df['is_service_row'] == False].copy().reset_index(drop=True)

            # ============================
            # 4. NUMERIC
            # ============================
            self.log.emit("🔢 Hesaplama чисел...")

            numeric_cols = ['arrival_room', 'arrival_paidpax', 'arrival_adult',
                            'arrival_paidchd', 'arrival_freechd', 'arrival_baby', 'night_room',
                            'night_paidpax', 'night_adult', 'night_paidchd', 'night_freechd',
                            'night_baby', 'local_revenue', 'eur_revenue', 'eur_rev.%',
                            'eur_avg_perroom', 'eur_avg_perpaidpax', 'avg_paidpax_night',
                            'avg_rm.night', 'r.occ_%', 'b.occ_%']

            for col in numeric_cols:
                if col in df_clean.columns:
                    df_clean[col] = pd.to_numeric(
                        df_clean[col].astype(str).str.replace(',', '.').str.replace(' ', ''),
                        errors='coerce'
                    )

            # ============================
            # 5. GROUP RULES
            # ============================
            self.log.emit("🏢 Agencta perfomans ...")

            agency_group_rules = {
                'Anex Tour': ['ANEX-'],
                'AKAY TOUR': ['AKAY-'],
                'ARELES (EUROPEHOL)': ['ARELES-'],
                'BEDSOPIA / PRIME TRAVEL': ['BEDSOPIA'],
                'BOOKING.COM': ['BOOKING.COM'],
                'COMP': ['GM', 'COMP 3', 'SALES', 'KONSER', 'ATG', 'PANDEMI'],
                'CORENDON': ['CORENDON'],
                'DESTINATION SERVICES': ['DESTINATION-'],
                'ETS': ['ETS'],
                'EUROPE HOLIDAY': ['EUHOLIDAY-'],
                'FIBULA TRAVEL': ['FIBULA-'],
                'FIT TURIZM': ['FIT HOL-', 'FIT'],
                'GROUP': ['GROUP-'],
                'HOTELBEDS': ['HOTELBEDS-'],
                'HOUSE USE': ['HOUSE USE'],
                'INDIVIDUAL': ['INDIVIDUAL-'],
                'ITS': ['ITS-'],
                'KALANIT TOUR': ['KALANIT-'],
                'KEYF TRAVEL': ['KEYF TRAVEL-', 'SUNQUEST-'],
                'KILIT GLOBAL': ['KILIT-'],
                'MEETING POINT': ['FTI-'],
                'MOTUS': ['MOTUS-'],
                'ODEON TOUR': ['ODEON-'],
                'PASSO TOUR': ['PASSO-'],
                'PENINSULA': ['PENINSULA-'],
                'PGM HOLIDAY': ['PGM HOLIDAY-'],
                'RUSTAR': ['RUSTAR'],
                'SETUR': ['SETUR'],
                'SONAR TOUR': ['SONAR-'],
                'SUMMER TOUR': ['SUMMER-'],
                'TATILBUDUR': ['TATILBUDUR'],
                'WEB': ['WEB-'],
                'ZEYDE TURIZM': ['ZEYDE TURIZM']
            }

            def map_group(a):
                au = str(a).upper()
                for g, patterns in agency_group_rules.items():
                    for p in patterns:
                        if au.startswith(p.upper()) or f" {p.upper()}" in au:
                            return g
                return "SORSAT"

            df_clean['agency_group'] = df_clean['agency'].apply(map_group);

            # ============================
            # 6. DROP EMPTY
            # ============================
            df_clean = df_clean.dropna(subset=['month', 'market'])

            # ============================
            # 7. GROUP
            # ============================
            self.log.emit("📊 toplama agenta agenta данных...")
            df_clean['YIL'] = '2026'

            group_cols = ['agency_group', 'YIL', 'month', 'market','agency']
            df_grouped = df_clean.groupby(group_cols, dropna=False)[numeric_cols].sum(min_count=1).reset_index()

            df_by_group = df_grouped.groupby(['month', 'agency_group']).agg({
                'arrival_room': 'sum',
                'night_room': 'sum',
                'eur_revenue': 'sum',
                'eur_avg_perpaidpax': 'mean'
            }).reset_index()

            df_by_market = df_grouped.groupby(['month', 'market']).agg({
                'arrival_room': 'sum',
                'night_room': 'sum',
                'eur_revenue': 'sum',
                'eur_avg_perpaidpax': 'mean'
            }).reset_index()

            # ============================
            # 8. SAVE EXCEL
            # ============================
            self.log.emit("💾 Yeni  Excel oluşturma...")

            def format_ws(ws):
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=11)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Header
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Data rows
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.border = border
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "#,##0.00"

                # Auto width
                for column in ws.columns:
                    max_len = 0
                    col_letter = get_column_letter(column[0].column)
                    for c in column:
                        if c.value:
                            max_len = max(max_len, len(str(c.value)))
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

                ws.freeze_panes = "A2"

            with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
                df_grouped.to_excel(writer, sheet_name="Summary", index=False)
                df_by_group.to_excel(writer, sheet_name="By Group", index=False)
                df_by_market.to_excel(writer, sheet_name="By Market", index=False)

                wb = writer.book

                for name in ["Summary", "By Group", "By Market"]:
                    ws = wb[name]
                    format_ws(ws)

            self.finished.emit(OUTPUT_PATH)

        except Exception as e:
            self.error.emit(str(e))
